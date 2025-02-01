
import requests
import os
import json
import subprocess
import logging
import time
import pandas as pd
from http.server import HTTPServer, BaseHTTPRequestHandler
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from dotenv import load_dotenv
from urllib.parse import urlencode
from datetime import datetime, timedelta
from selenium.webdriver.chrome.service import Service
from threading import Event
from http.server import HTTPServer, BaseHTTPRequestHandler
from openpyxl import load_workbook


# Load environment variables
load_dotenv()
auth_code_event = Event()

# Global Configuration Section
CONFIG_FILE = "config.json"
with open(CONFIG_FILE, "r") as f:
    config = json.load(f)

API_BASE_URL = "https://webexapis.com/v1"  # Base URL for Webex APIs
API_BASE_URL_EU = "https://api.wxcc-eu2.cisco.com" 
AUTH_METHOD = config.get("auth_method", "curl")  # Default to cURL
#API_BASE_URL = "https://api.wxcc-eu1.cisco.com"
CLIENT_ID = "C981dab8ee0604bbb01f905030bcb13a7fea0fde4a9a166c5c4f508a40bef99af"  # Your Webex App's Client ID
CLIENT_SECRET = "7310096a885b79a3094fe1e5efaa66377ec4f48a63d7502dce1e8d836931916c"  # Your Webex App's Client Secret
REDIRECT_URI = "http://localhost:8089/callback"  # The redirect URI registered in your Webex app
ORG_ID = "d5b25589-1c50-4233-b21e-ec4cb78ee781"
AUTH_CODE = "MWE1NTQ4MzYtZmM4NS00ZDJjLWE0ZDMtNWEzZmIyM2ZhNjUzMjI4OGZmOTEtZmFi_PE93_d5b25589-1c50-4233-b21e-ec4cb78ee781"  # The authorization code received after user authentication
LOG_FILE = "webex_auth.log"  # Log file path
CHANNEL_TYPE = "chat"  # Specify the channel type (e.g., "chat")
TOKEN_FILE = "access_token.json"
SCOPE = (
    "spark:kms cloud-contact-center:pod_conv cjp:user "
    "spark:people_read cjp:config cjp:config_read cjds:admin_org_read"
)



# Logging Configuration
logging.basicConfig(
    level=logging.DEBUG,  # Change to DEBUG for detailed logs
    format="%(asctime)s - %(levelname)s - %(funcName)s - %(lineno)d - %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE),  # Logs to a file
        logging.StreamHandler()  # Logs to the console
    ]
)

class AuthCodeHandler(BaseHTTPRequestHandler):
    """
    HTTP handler to process the redirect and capture the authorization code.
    """
    def do_GET(self):
        if "/callback" in self.path:
            query = self.path.split("?")[1]
            params = dict(param.split("=") for param in query.split("&"))
            auth_code = params.get("code")
            logging.info(f"Authorization code received: {auth_code}")

            # Send response to browser
            self.send_response(200)
            self.end_headers()
            self.wfile.write(b"Authorization successful. You can close this window.")

            # Save the auth code in the server object
            self.server.auth_code = auth_code

            # Signal that the auth code has been captured
            auth_code_event.set()
            self.server.shutdown()

def wait_for_authorization_code(driver, timeout=60):
    """
    Wait for the browser to redirect to the callback URL with the authorization code.

    Args:
        driver (WebDriver): Selenium WebDriver instance.
        timeout (int): Maximum time to wait for the authorization code (in seconds).

    Returns:
        str: The authorization code from the callback URL.
    """
    start_time = time.time()
    while time.time() - start_time < timeout:
        # Get the current URL
        current_url = driver.current_url
        if "code=" in current_url and "state=" in current_url:
            # Extract the authorization code
            from urllib.parse import urlparse, parse_qs

            query_params = parse_qs(urlparse(current_url).query)
            auth_code = query_params.get("code", [None])[0]
            if auth_code:
                print(f"Authorization code received: {auth_code}")
                return auth_code
        time.sleep(1)  # Wait a short period before checking again

    raise TimeoutError("Timeout waiting for authorization code.")

    # Example usage in your OAuth flow
    try:
        auth_url = "YOUR_AUTH_URL_HERE"
        driver.get(auth_url)
        print("Waiting for user to complete sign-in...")
        
        # Wait for the authorization code
        authorization_code = wait_for_authorization_code(driver)
        print(f"Authorization Code: {authorization_code}")
        
        # Close the browser
        driver.quit()

    except TimeoutError as e:
        print(f"Error: {e}")
        driver.quit()

def start_local_server():
    """
    Start a local HTTP server to capture the authorization code.
    """
    server = HTTPServer(("localhost", 8089), AuthCodeHandler)
    try:
        logging.info("Starting local server to capture the authorization code...")
        server.serve_forever()
    finally:
        server.server_close()
        logging.info("Local server has been shut down.")
    return getattr(server, "auth_code", None)

def curl_auth(auth_url):
    """
    Use cURL to interact with the Webex authorization URL.
    """
    try:
        logging.info("Starting cURL-based authorization session...")
        curl_command = f'curl -L "{auth_url}"'
        result = subprocess.run(curl_command, shell=True, capture_output=True, text=True)
        
        if result.returncode == 0:
            logging.info("cURL authorization session completed.")
            logging.debug(f"cURL Output: {result.stdout}")
        else:
            logging.error(f"cURL failed with return code {result.returncode}")
            logging.error(f"cURL Error: {result.stderr}")
    except Exception as e:
        logging.exception(f"Failed to complete cURL authorization: {e}")

def browser_auth(auth_url):
    """
    Open an anonymous browser session to the Webex authorization URL and wait for the user to complete the process.

    Args:
        auth_url (str): The Webex OAuth authorization URL.
    """
    try:
        # Configure Chrome options
        chrome_options = Options()
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--no-sandbox")

        logging.info("Opening anonymous browser session...")
        driver_service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=driver_service, options=chrome_options)
        driver.get(auth_url)

        logging.info("Waiting for the user to sign in and authorize the app...")

        # Wait for the authorization code
        authorization_code = wait_for_authorization_code(driver)
        logging.info(f"Authorization process completed. Code: {authorization_code}")
        driver.quit()
        return authorization_code
    except TimeoutError:
        logging.error("Timeout waiting for user to complete sign-in.")
        driver.quit()
        return None
    except Exception as e:
        logging.exception(f"Failed to complete browser authorization: {e}")
        driver.quit()
        return None

def get_auth_url():
    """
    Generate the Webex OAuth Authorization URL with proper encoding.

    Returns:
        str: The properly encoded authorization URL.
    """
    params = {
        "client_id": CLIENT_ID,
        "response_type": "code",
        "redirect_uri": REDIRECT_URI,
        "scope": SCOPE,
        "state": "xyz123",
    }
    auth_url = f"https://webexapis.com/v1/authorize?{urlencode(params)}"
    logging.debug(f"Generated Authorization URL: {auth_url}")
    return auth_url

def get_access_token(auth_code):
    """
    Exchange the authorization code for an access token.

    Args:
        auth_code (str): The authorization code received after user authentication.

    Returns:
        dict: Access token and expiration information if successful, None otherwise.
    """
    url = f"{API_BASE_URL}/access_token"
    headers = {
        "Content-Type": "application/x-www-form-urlencoded",
    }
    data = {
        "grant_type": "authorization_code",
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "code": auth_code,
        "redirect_uri": REDIRECT_URI,
    }

    try:
        logging.info("Requesting access token...")
        response = requests.post(url, headers=headers, data=data)

        if response.status_code == 200:
            token_data = response.json()
            # Add expiration time (current time + expires_in seconds)
            token_data["expires_at"] = (
                datetime.utcnow() + timedelta(seconds=token_data["expires_in"])
            ).strftime("%Y-%m-%dT%H:%M:%SZ")
            logging.info("Access token successfully retrieved.")
            return token_data
        else:
            logging.error(f"Failed to retrieve access token: {response.status_code} - {response.text}")
            return None
    except requests.exceptions.RequestException as e:
        logging.exception(f"Error while requesting access token: {e}")
        return None

def save_token(token_data):
    """
    Save the access token and its expiration time to a file.

    Args:
        token_data (dict): Token data containing access_token and expires_at.
    """
    try:
        with open(TOKEN_FILE, "w") as file:
            json.dump(token_data, file, indent=4)
        logging.info(f"Access token saved to {TOKEN_FILE}.")
    except Exception as e:
        logging.exception(f"Failed to save access token: {e}")

def load_token():
    """
    Load the access token from the file if it exists and is valid.

    Returns:
        str: The access token if valid, None otherwise.
    """
    if not os.path.exists(TOKEN_FILE):
        return None

    try:
        with open(TOKEN_FILE, "r") as file:
            token_data = json.load(file)
            expires_at = datetime.strptime(token_data["expires_at"], "%Y-%m-%dT%H:%M:%SZ")
            if datetime.utcnow() < expires_at:
                logging.info("Access token is valid.")
                return token_data["access_token"]
            else:
                logging.info("Access token has expired.")
                return None
    except Exception as e:
        logging.exception(f"Failed to load access token: {e}")
        return None

def search_tasks(access_token, org_id, start_time, end_time, queue_name):
    """
    Search for live chat tasks using the Webex Search API with a GraphQL query.

    Args:
        access_token (str): The access token for authentication.
        org_id (str): The organization ID.
        start_time (int): Start time in epoch milliseconds.
        end_time (int): End time in epoch milliseconds.
        queue_name (str): The name of the queue to filter tasks.

    Returns:
        list: List of filtered tasks if successful, None otherwise.
    """
    url = f"https://api.wxcc-eu2.cisco.com/search?orgId={org_id}"
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json",
        "Accept": "application/json",
    }
    query = """
    query($startTime: Long!, $endTime: Long!) {
        task(
            from: $startTime,
            to: $endTime,
            filter: {
                channelType: { equals: chat }
            },
            pagination: { cursor: "0" }
        ) {
            tasks {
                id
                channelType
                createdTime
                endedTime
                captureRequested
                isActive
                status
                
            }
        }
    }
    """
    payload = {
        "query": query,
        "variables": {
            "startTime": start_time,
            "endTime": end_time,
        },
    }

    try:
        logging.debug(f"Headers: {headers}")
        logging.debug(f"Payload: {json.dumps(payload, indent=4)}")

        response = requests.post(url, headers=headers, json=payload)

        if response.status_code == 200:
            results = response.json()
            logging.info(f"Search successful. Found {len(results.get('data', {}).get('task', {}).get('tasks', []))} tasks.")
            logging.debug(f"API Response: {json.dumps(results, indent=4)}")

            # Filter tasks by queue name in Python
            return results.get('data', {}).get('task', {}).get('tasks', [])
        else:
            logging.error(f"Failed to search tasks: {response.status_code} - {response.text}")
            return None
    except requests.exceptions.RequestException as e:
        logging.exception(f"An error occurred while searching: {e}")
        return None
    
def get_transcript(access_token, org_id, task_id):
    """
    Retrieve the transcript or capture details for a specific task using the captures/query API.

    Args:
        access_token (str): The access token for authentication.
        org_id (str): The organization ID.
        task_id (str): The task ID.

    Returns:
        dict: Capture details if successful, None otherwise.
    """
    url = "https://api.wxcc-eu2.cisco.com/v1/captures/query"
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json",
        "Accept": "application/json",
    }
    payload = {
        "query": {
            "orgId": org_id,
            "taskIds": [task_id],
            "urlExpiration": 3600,  # Set URL expiration to 1 hour
            "includeSegments": True
        }
    }

    try:
        logging.info(f"Retrieving capture for task ID: {task_id}")
        response = requests.post(url, headers=headers, json=payload)
        if response.status_code == 200:
            capture_data = response.json()
            file_path = capture_data["data"][0]["transcription"][0]["filePath"]
            logging.info(f"Transcript available at: {file_path}")
            logging.debug(f"Capture API Response: {json.dumps(capture_data, indent=4)}")
            if file_path:
                    process_transcript_to_excel(task_id, file_path)
            else:
                    logging.error(f"No URL found to download transcript for task {task_id}")
            return capture_data
        elif response.status_code == 404:
            logging.error(f"Task ID {task_id} not found or captures not available.")
        else:
            logging.error(f"Failed to retrieve capture: {response.status_code} - {response.text}")
    except requests.exceptions.RequestException as e:
        logging.exception(f"An error occurred while retrieving the capture: {e}")

    return None

def generate_unique_filename(base_name="transcripts", extension=".xlsx"):
    """
    Generate a unique filename by appending a timestamp.

    Args:
        base_name (str): The base name of the file.
        extension (str): The file extension.

    Returns:
        str: A unique filename with a timestamp.
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{base_name}_{timestamp}{extension}"

def process_transcript_to_excel(task_id, url, excel_file="transcripts.xlsx"):
    """
    Fetch JSON transcript data from a URL and write it into an Excel file, associating it with a task ID.

    Args:
        task_id (str): The task ID associated with the transcript.
        url (str): URL containing JSON transcript data.
        excel_file (str): Path to the output Excel file.
    """
    try:
        # Fetch the JSON data from the URL
        response = requests.get(url)
        if response.status_code == 200:
            json_data = response.json()

            # Flatten the JSON into a tabular format
            rows = []
            for entry in json_data:
                flattened_entry = {
                    "taskId": task_id,  # Associate the task ID with each row
                    "id": entry.get("id"),
                    "aliasId": entry.get("aliasId"),
                    "direction": entry.get("direction"),
                    "message": entry.get("message"),
                    "timestamp": entry.get("timestamp"),
                    "participant_name": entry.get("participant", {}).get("name"),
                    "participant_role": entry.get("participant", {}).get("role"),
                    "participant_userId": entry.get("participant", {}).get("userId"),
                    "redacted": entry.get("redacted"),
                }
                rows.append(flattened_entry)

            # Create a DataFrame
            df = pd.DataFrame(rows)

            # Check if the Excel file already exists
            if os.path.exists(excel_file):
                with pd.ExcelWriter(excel_file, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
                    sheet_name = "Transcripts"
                    workbook = load_workbook(excel_file)
                    if sheet_name in workbook.sheetnames:
                        # Append to the existing sheet
                        startrow = writer.sheets[sheet_name].max_row
                        df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=startrow, header=False)
                    else:
                        # Create a new sheet
                        df.to_excel(writer, index=False, sheet_name=sheet_name)
            else:
                # Create a new Excel file with the data
                with pd.ExcelWriter(excel_file, mode="w", engine="openpyxl") as writer:
                    df.to_excel(writer, index=False, sheet_name="Transcripts")

            print(f"Processed data for task ID {task_id} from URL: {url}. Excel file: {excel_file}")
        else:
            print(f"Failed to fetch data from {url}. Status code: {response.status_code}")
    except Exception as e:
        print(f"Error processing task ID {task_id} from URL {url}: {e}")




if __name__ == "__main__":
    logging.info("Starting Webex OAuth process.")

    # Load the access token if it exists and is valid
    access_token = load_token()

    if not access_token:
        logging.info("No valid access token found. Starting authorization process...")

        # Start the local server in a separate thread
        from threading import Thread
        server_thread = Thread(target=start_local_server, daemon=True)
        server_thread.start()

        # Generate the authorization URL
        auth_url = get_auth_url()

        # Open the browser for user sign-in
        browser_auth(auth_url)

        # Wait for the authorization code to be captured
        logging.info("Waiting for the authorization code...")
        auth_code_event.wait()  # Block until the event is set

        # Retrieve the authorization code from the server
        auth_code = getattr(start_local_server, "auth_code", None)

        if auth_code:
            logging.info(f"Captured Authorization Code: {auth_code}")

            # Exchange the authorization code for an access token
            token_data = get_access_token(auth_code)

            if token_data:
                # Save the token to a file for future use
                save_token(token_data)
                access_token = token_data.get("access_token")
            else:
                logging.error("Failed to retrieve access token.")
                exit(1)
        else:
            logging.error("Failed to capture the authorization code.")
            exit(1)
    else:
        logging.info("Using the existing valid access token.")

    # Define the time range for live chats (last 24 hours)
    start_time = int((time.time() - (4 * 7 * 24 * 60 * 60)) * 1000)  # 24 hours ago in epoch milliseconds
    end_time = int(time.time() * 1000)  # Current time in epoch milliseconds
    queueName = "ShowCase001_LiveChat_RijksVideo"

    # Step 1: Search for live chats
    logging.info(f"Starting searching for tasks between {start_time} and {end_time}")
    tasks = search_tasks(access_token, ORG_ID, start_time, end_time,queueName)

    # Step 2: Retrieve and log transcripts
    try:
        for task in tasks:
            task_id = task.get("id")
            
            if task_id:
                transcript = get_transcript(access_token, ORG_ID,task_id)
                if transcript:
                    logging.info(f"Transcript for task ID {task_id}: {transcript}")
    except Exception as e:
        logging.exception(f"Failed processing tasks: {e}")